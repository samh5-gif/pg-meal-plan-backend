"""
Project GAIN -- Meal Plan PDF + Recipe Guide Generator
Backend: Flask API with Anthropic AI for recipe generation
"""

import os, re, io, base64, json, tempfile
from flask import Flask, request, jsonify
from flask_cors import CORS
import openpyxl
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4
from reportlab.pdfbase.pdfmetrics import stringWidth
import anthropic

from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont

def _register_fonts():
    possible_dirs = [
        os.path.join(os.path.dirname(os.path.abspath(__file__)), 'fonts'),
        os.path.join(os.getcwd(), 'fonts'),
        '/opt/render/project/src/fonts',
    ]
    for font_dir in possible_dirs:
        try:
            r = os.path.join(font_dir, 'Inter-Regular.ttf')
            b = os.path.join(font_dir, 'Inter-Bold.ttf')
            xb = os.path.join(font_dir, 'Inter-ExtraBold.ttf')
            if not all(os.path.exists(f) for f in [r, b, xb]):
                continue
            pdfmetrics.registerFont(TTFont('Inter', r))
            pdfmetrics.registerFont(TTFont('Inter-Bold', b))
            pdfmetrics.registerFont(TTFont('Inter-ExtraBold', xb))
            pdfmetrics.registerFontFamily('Inter', normal='Inter', bold='Inter-Bold')
            return True
        except Exception:
            continue
    return False

INTER_AVAILABLE = _register_fonts()

app = Flask(__name__)
CORS(app, resources={r"/*": {"origins": "*"}})

@app.after_request
def add_cors(response):
    response.headers['Access-Control-Allow-Origin'] = '*'
    response.headers['Access-Control-Allow-Headers'] = 'Content-Type'
    response.headers['Access-Control-Allow-Methods'] = 'GET, POST, OPTIONS'
    return response

@app.route('/upload', methods=['OPTIONS'])
@app.route('/clarify', methods=['OPTIONS'])
@app.route('/generate-pdf', methods=['OPTIONS'])
@app.route('/health', methods=['OPTIONS'])
def options():
    return '', 204

def get_claude():
    return anthropic.Anthropic(api_key=os.environ.get('ANTHROPIC_API_KEY', ''))

_sessions = {}

# ---------------------------------------------------------------------------
# Ingredient note logic
# ---------------------------------------------------------------------------
# These ingredients should show "(raw weight)" after their name in the PDF
# because coaches weigh them raw but clients cook them.
RAW_WEIGHT_KEYWORDS = [
    'chicken', 'turkey', 'beef', 'mince', 'steak', 'salmon', 'cod', 'tuna',
    'tilapia', 'haddock', 'trout', 'sea bass', 'mackerel', 'prawn', 'shrimp',
    'pork', 'lamb', 'venison', 'duck', 'quorn', 'tofu',
]

# These ingredients should show "(uncooked weight)" after their name.
UNCOOKED_WEIGHT_KEYWORDS = [
    'pasta', 'rice', 'oats', 'porridge oats', 'noodles', 'quinoa',
    'couscous', 'bulgur', 'lentil', 'chickpea',
]

def get_ingredient_note(food_name):
    """
    Returns a note string like '(raw weight)' or '(uncooked weight)' if the
    ingredient warrants one. Returns empty string otherwise.
    """
    name = food_name.lower()
    for kw in RAW_WEIGHT_KEYWORDS:
        if kw in name:
            return '(raw weight)'
    for kw in UNCOOKED_WEIGHT_KEYWORDS:
        if kw in name:
            return '(uncooked weight)'
    return ''

# ---------------------------------------------------------------------------
# Ambiguous count/size detection (eggs, fruit, veg sold by unit not weight)
# ---------------------------------------------------------------------------
# Foods excluded from the whole-unit detection entirely (measured by weight
# as normal, no size hint needed).
EXCLUSIONS = ['egg white', 'protein water', 'egg powder', 'passionfruit']

# Only flag foods that are genuinely sold/used as whole units and where
# cutting to an exact gram weight is impractical:
#   - Eggs: you use a whole egg, you can't cut one to 60g
#   - Fruit: unlikely to cut part of a piece and store the rest; the whole
#     fruit is consumed in one go or very soon after
# Deliberately excluded:
#   - Avocado: high fat, accurate gram weight is critical for macros
#   - Root veg (carrot, potato, sweet potato): easy to cut to weight
#   - Onion, courgette: easy to cut to weight
AMBIGUOUS_PATTERNS = [
    # Eggs
    r'\bboiled eggs?\b',
    r'\bscrambled eggs?\b',
    r'\bfried eggs?\b',
    r'\bpoached eggs?\b',
    r'^eggs?$',
    r'\beggs?\b',
    # Fruit
    r'\bapple\b',
    r'\bbanana\b',
    r'\bpeach\b',
    r'\bplum\b',
    r'\borange\b',
    r'\bpear\b',
    r'\bmango\b',
    r'\bkiwi\b',
    r'\bgrapes?\b',
    r'\bblueberr',
    r'\bstrawberr',
    r'\braspberr',
    r'\bblackberr',
    r'\bmelon\b',
    r'\bpineapple\b',
    r'\bcherry\b',
    r'\bfig\b',
    r'\bdate\b',
    r'\bapricot\b',
    r'\bnectarine\b',
    r'\bclementine\b',
    r'\btangerine\b',
    r'\bgrapefruit\b',
]

def check_ambiguous(qty_g, food_name):
    name_lower = food_name.lower()
    for excl in EXCLUSIONS:
        if excl in name_lower:
            return None
    for pattern in AMBIGUOUS_PATTERNS:
        if re.search(pattern, name_lower, re.IGNORECASE):
            return {
                'food': food_name,
                'qty_g': qty_g,
                'key': name_lower.strip(),
                'suggestion': None,
            }
    return None


# Size reference data -- used to generate a practical count hint shown in the
# coach-facing clarification screen only, NOT used to replace the quantity in
# the PDF. The PDF always shows the exact gram weight from the Excel.
# Size reference data for whole-unit foods only.
# Veg (carrot, potato, sweet potato, onion, courgette) and avocado are
# intentionally excluded -- coaches should weigh these accurately.
#
# Keys are matched with whole-word regex (r'\bKEY\b') so that e.g. 'apple'
# does NOT match 'pineapple'. More specific entries (pineapple, grapefruit)
# must be listed before shorter overlapping keys (apple, grape) so the first
# match wins correctly.
_SIZE_DATA = [
    ('pineapple',   r'\bpineapple\b',  160, 'slice of pineapple',  'a standard pineapple slice weighs around 100-200g; a small whole pineapple around 800g'),
    ('grapefruit',  r'\bgrapefruit\b', 300, 'grapefruit',          'a whole grapefruit weighs around 280-320g'),
    ('clementine',  r'\bclementine\b', 85,  'clementine',          'a clementine weighs around 80-90g'),
    ('tangerine',   r'\btangerine\b',  85,  'tangerine',           'a tangerine weighs around 80-90g'),
    ('nectarine',   r'\bnectarine\b',  150, 'nectarine',           'a nectarine weighs around 140-160g'),
    ('apricot',     r'\bapricot\b',    55,  'apricot',             'an apricot weighs around 50-60g'),
    ('melon',       r'\bmelon\b',      200, 'slice of melon',      'a standard melon slice weighs around 150-200g'),
    ('egg',         r'\beggs?\b',      63,  'large egg',           'UK large eggs weigh 63-73g each'),
    ('apple',       r'\bapple\b',      182, 'large apple',         'large apples weigh around 180-220g'),
    ('banana',      r'\bbanana\b',     120, 'medium banana',       'medium bananas weigh around 110-130g'),
    ('peach',       r'\bpeach\b',      175, 'large peach',         'large peaches weigh around 170-200g'),
    ('plum',        r'\bplum\b',       70,  'large plum',          'large plums weigh around 65-85g'),
    ('orange',      r'\borange\b',     180, 'large orange',        'large oranges weigh around 175-210g'),
    ('pear',        r'\bpear\b',       170, 'medium pear',         'medium pears weigh around 160-190g'),
    ('mango',       r'\bmango\b',      350, 'whole mango',         'a whole large mango weighs around 300-400g'),
    ('kiwi',        r'\bkiwi\b',       70,  'kiwi',                'a standard kiwi weighs around 70g'),
    ('grape',       r'\bgrapes?\b',    5,   'grape',               'individual grapes weigh around 5-8g'),
    ('fig',         r'\bfig\b',        60,  'fig',                 'a fresh fig weighs around 55-65g'),
]

_CACHE_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'size_cache.json')

def _load_cache():
    try:
        with open(_CACHE_FILE) as f:
            return json.load(f)
    except Exception:
        return {}

def _save_cache(cache):
    try:
        with open(_CACHE_FILE, 'w') as f:
            json.dump(cache, f, indent=2)
    except Exception:
        pass

_size_cache = _load_cache()


def lookup_size_suggestions(ambiguous_items):
    """
    Builds two fields for each ambiguous item:
      - 'suggestion': the full informational string shown in the UI hint box
                      e.g. "100g = 1-2 large eggs (UK large eggs weigh 63-73g each)"
      - 'short':      the concise count/size string pre-filled into the input box
                      e.g. "1-2 large eggs"
    Neither field is used by the PDF. The PDF always renders the exact gram
    weight from the Excel file.
    """
    for item in ambiguous_items:
        name_lower = item['food'].lower()
        qty = item['qty_g']
        qty_int = int(qty) if qty == int(qty) else qty

        matched = None
        for (_key, pattern, unit_g, unit_name, reference) in _SIZE_DATA:
            if re.search(pattern, name_lower, re.IGNORECASE):
                matched = (unit_g, unit_name, reference)
                break

        if matched:
            unit_g, unit_name, reference = matched
            count = qty / unit_g
            if count <= 0.4:
                short = f"a small portion ({qty_int}g)"
            elif count <= 0.65:
                article = "an" if unit_name[0] in "aeiou" else "a"
                short = f"half {article} {unit_name}"
            elif count <= 1.3:
                short = f"1 {unit_name}"
            elif count <= 1.7:
                short = f"1-2 {unit_name}s"
            else:
                n = round(count)
                short = f"{n} {unit_name}s"

            item['suggestion'] = f"{qty_int}g = {short} ({reference})"
            item['short'] = short
        else:
            item['suggestion'] = None
            item['short'] = ''

    unknown = [i for i in ambiguous_items if i['suggestion'] is None]

    if unknown:
        still_unknown = []
        for item in unknown:
            cache_key = f"{item['food'].lower().strip()}:{int(item['qty_g']) if item['qty_g']==int(item['qty_g']) else item['qty_g']}"
            if cache_key in _size_cache:
                cached = _size_cache[cache_key]
                item['suggestion'] = cached.get('suggestion', '')
                item['short']      = cached.get('short', '')
            else:
                still_unknown.append(item)

        if still_unknown:
            try:
                items_text = '\n'.join(
                    f"- {i['food']}: {int(i['qty_g']) if i['qty_g']==int(i['qty_g']) else i['qty_g']}g"
                    for i in still_unknown
                )
                prompt = (
                    "You are a nutrition expert. For each ingredient and gram weight below, "
                    "give the most accurate practical size equivalent.\n"
                    "Provide two fields:\n"
                    "  'suggestion': full string e.g. '50g = 1 small egg (large eggs weigh 63-73g)'\n"
                    "  'short': concise count/size only e.g. '1 small egg'\n\n"
                    "Respond ONLY as valid JSON:\n"
                    '{"suggestions": [{"food": "X", "suggestion": "50g = 1 small egg (large eggs weigh 63-73g)", "short": "1 small egg"}]}\n\n'
                    f"Ingredients:\n{items_text}"
                )
                resp = get_claude().messages.create(
                    model='claude-haiku-4-5-20251001',
                    max_tokens=400,
                    messages=[{'role': 'user', 'content': prompt}]
                )
                raw = resp.content[0].text.strip()
                raw = re.sub(r'^```json\s*', '', raw, flags=re.MULTILINE)
                raw = re.sub(r'^```\s*', '', raw, flags=re.MULTILINE)
                raw = re.sub(r'\s*```$', '', raw)
                data = json.loads(raw)

                sugg_map = {}
                for s in data.get('suggestions', []):
                    k = s['food'].lower().strip()
                    sugg_map[k] = s
                    sugg_map[re.sub(r'\s*\([^)]+\)', '', k).strip()] = s

                cache_updated = False
                for item in still_unknown:
                    k  = item['food'].lower().strip()
                    ck = re.sub(r'\s*\([^)]+\)', '', k).strip()
                    match = sugg_map.get(k) or sugg_map.get(ck)
                    qty_int = int(item['qty_g']) if item['qty_g']==int(item['qty_g']) else item['qty_g']
                    if match:
                        item['suggestion'] = match.get('suggestion', '')
                        item['short']      = match.get('short', '')
                        cache_key = f"{k}:{qty_int}"
                        _size_cache[cache_key] = {'suggestion': item['suggestion'], 'short': item['short']}
                        cache_updated = True
                    else:
                        item['suggestion'] = f"{qty_int}g -- please confirm the practical size"
                        item['short'] = ''

                if cache_updated:
                    _save_cache(_size_cache)

            except Exception:
                for item in still_unknown:
                    qty_int = int(item['qty_g']) if item['qty_g']==int(item['qty_g']) else item['qty_g']
                    item['suggestion'] = f"{qty_int}g -- please confirm the practical size"
                    item['short'] = ''

    return ambiguous_items


def is_fruit(food_name):
    fruits = {'apple','banana','peach','plum','orange','pear','mango','kiwi',
              'grape','blueberr','strawberr','raspberr','blackberr','melon',
              'pineapple','cherry','lemon','lime','grapefruit','fig','date',
              'apricot','nectarine','pomegranate'}
    name = food_name.lower()
    return any(f in name for f in fruits)


# ---------------------------------------------------------------------------
# Excel parsing
# ---------------------------------------------------------------------------
def parse_excel(file_bytes, filename):
    stem = os.path.splitext(filename)[0]
    stem = re.sub(r'(?i)meal[\s_-]*plan[\s_-]*[-_]*', '', stem)
    stem = re.sub(r'[\s_-]+\d+[\s_-]*$', '', stem)
    stem = re.sub(r'[_]+', ' ', stem).strip(' -_()[]')
    client_name = ' '.join(w.capitalize() for w in stem.split() if w) or 'Client'

    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    days = []
    day_num = 0

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        nl = sheet_name.lower()
        if any(skip in nl for skip in ['shopping', 'foodlist', 'food list', 'targets', 'summary', 'notes']):
            continue
        has_day_indicator = any(w in nl for w in ['meal', 'day', 'week'])

        day_num += 1
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue

        header_row = None
        for i, row in enumerate(rows):
            row_s = [str(c).lower() if c else '' for c in row]
            if any('meal' in c for c in row_s) and any('food' in c or 'item' in c for c in row_s):
                header_row = i
                break
        if header_row is None:
            continue

        if not has_day_indicator:
            has_meal_data = False
            for row in rows[header_row+1:header_row+5]:
                if row and row[0] and str(row[0]).strip().lower() not in ('none', 'nan', ''):
                    has_meal_data = True
                    break
            if not has_meal_data:
                continue

        headers = [str(c).lower().strip() if c else '' for c in rows[header_row]]

        def col(kws):
            for i, h in enumerate(headers):
                if any(k in h for k in kws):
                    return i
            return None

        meal_col = col(['meal'])
        food_col = col(['food', 'item', 'ingredient'])
        qty_col  = col(['quantity', 'qty', 'amount', 'weight'])
        kcal_col = col(['calori', 'kcal', 'energy'])
        prot_col = col(['protein'])
        fat_col  = col(['fat'])
        carb_col = col(['carb'])

        def to_f(v):
            try: return float(v)
            except: return 0.0

        meals_dict = {}
        for row in rows[header_row+1:]:
            if not any(row):
                continue
            mv = str(row[meal_col]).strip() if meal_col is not None and row[meal_col] else ''
            if not mv or mv.lower() in ('none', 'nan', 'meal', 'meal name', 'meal label'):
                continue
            if mv.lower() in ('total', 'daily total', 'totals'):
                continue
            fv = str(row[food_col]).strip() if food_col is not None and row[food_col] else ''
            if not fv or fv.lower() == 'nan':
                continue

            qty_g  = to_f(row[qty_col]  if qty_col  is not None else 0)
            kcal_v = to_f(row[kcal_col] if kcal_col is not None else 0)
            prot_v = to_f(row[prot_col] if prot_col is not None else 0)
            fat_v  = to_f(row[fat_col]  if fat_col  is not None else 0)
            carb_v = to_f(row[carb_col] if carb_col is not None else 0)

            if mv not in meals_dict:
                meals_dict[mv] = {'meal_label': mv, 'kcal': 0, 'prot': 0, 'fat': 0, 'carb': 0, 'ingredients': []}

            meals_dict[mv]['kcal'] += kcal_v
            meals_dict[mv]['prot'] += prot_v
            meals_dict[mv]['fat']  += fat_v
            meals_dict[mv]['carb'] += carb_v
            meals_dict[mv]['ingredients'].append({
                'food': fv,
                'qty_g': qty_g,
                'excel_qty_g': qty_g,
            })

        def meal_key(m):
            nums = re.findall(r'\d+', m)
            return int(nums[0]) if nums else 0

        meals = sorted(meals_dict.values(), key=lambda m: meal_key(m['meal_label']))
        for m in meals:
            m['kcal'] = round(m['kcal'])
            m['prot'] = round(m['prot'], 1)
            m['fat']  = round(m['fat'],  1)
            m['carb'] = round(m['carb'], 1)
            m['dish_name'] = ''
            m['recipe_steps'] = []

        days.append({
            'day_num': day_num,
            'sheet': sheet_name,
            'total_kcal': round(sum(m['kcal'] for m in meals)),
            'total_prot': round(sum(m['prot'] for m in meals), 1),
            'total_fat':  round(sum(m['fat']  for m in meals), 1),
            'total_carb': round(sum(m['carb'] for m in meals), 1),
            'meals': meals,
        })

    shopping_lists = parse_shopping_lists(wb)
    return client_name, days, shopping_lists


def parse_shopping_lists(wb):
    shopping_lists = []
    for sheet_name in wb.sheetnames:
        nl = sheet_name.lower()
        if 'shopping' not in nl:
            continue
        ws = wb[sheet_name]
        rows = [r for r in ws.iter_rows(values_only=True) if any(r)]
        if not rows:
            continue
        header = rows[0]
        col2 = str(header[1]) if len(header) > 1 and header[1] else ''
        note = ''
        if '*' in col2:
            parts = col2.split('*')
            note = parts[1].strip().strip('*').strip() if len(parts) > 1 else ''
        if not note:
            note = 'assumes each meal prepared once'
        items = []
        for row in rows[1:]:
            food = str(row[0]).strip() if row[0] else ''
            qty  = row[1] if len(row) > 1 and row[1] is not None else ''
            if not food or food.lower() in ('none', 'nan', 'food item'):
                continue
            try:
                qty_val = float(qty)
                qty_str = f"{int(qty_val)}g" if qty_val == int(qty_val) else f"{qty_val}g"
            except:
                qty_str = str(qty) if qty else ''
            items.append({'food': food, 'qty': qty_str})
        if items:
            shopping_lists.append({'name': sheet_name, 'note': note, 'items': items})
    return shopping_lists


def find_ambiguous(days):
    seen, result = set(), []
    for day in days:
        for meal in day['meals']:
            for ing in meal['ingredients']:
                a = check_ambiguous(ing['qty_g'], ing['food'])
                if a and a['key'] not in seen:
                    seen.add(a['key'])
                    result.append(a)
    return result


def has_fruit(days):
    return any(is_fruit(ing['food']) for day in days for meal in day['meals'] for ing in meal['ingredients'])

def day_has_fruit(day):
    return any(is_fruit(ing['food']) for meal in day['meals'] for ing in meal['ingredients'])


# ---------------------------------------------------------------------------
# AI Recipe Generation
# ---------------------------------------------------------------------------
def _clean_json(raw):
    raw = raw.strip()
    raw = re.sub(r'^```json\s*', '', raw, flags=re.MULTILINE)
    raw = re.sub(r'^```\s*', '', raw, flags=re.MULTILINE)
    raw = re.sub(r'\s*```$', '', raw)
    m = re.search(r'\{.*\}', raw, re.DOTALL)
    if m:
        raw = m.group(0)
    for old, new in [(u'\u2018',"'"),(u'\u2019',"'"),(u'\u201c','"'),(u'\u201d','"'),
                     (u'\u2013','-'),(u'\u2014','-'),(u'\u2026','...')]:
        raw = raw.replace(old, new)
    try:
        return json.loads(raw)
    except json.JSONDecodeError:
        raw = raw.encode('ascii', 'replace').decode('ascii')
        return json.loads(raw)


def _generate_batch(claude, batch_meals, batch_num):
    meals_text = []
    for day_num, meal_label, ings_text in batch_meals:
        meals_text.append(f"Day {day_num} -- {meal_label}:\n{ings_text}")

    prompt = """You are a nutrition coach writing a recipe guide for a fitness client.

For each meal, generate a dish name and step-by-step cooking instructions.

Respond ONLY with valid JSON -- no markdown, no code fences:
{"meals": [{"day": 1, "meal_label": "Breakfast", "dish_name": "Creamy Porridge", "steps": ["Fill a medium saucepan with water and bring to a boil.", "Add the oats and stir continuously for 3 minutes."], "unclear": false, "unclear_reason": ""}]}

Rules:
- Steps should be practical and motivating, written directly to the client
- Keep dish names short and punchy -- maximum 5 words (e.g. "Scrambled Eggs on Sourdough" not "Scrambled Eggs and Egg White on Sourdough with Avocado and Spinach")
- Single-item meals (protein water, bar, fruit) need only 1-2 steps
- Set unclear to true only if you genuinely cannot determine the dish
- Never invent ingredients not in the list
- No nutrition advice in steps
- Use only plain ASCII characters -- no smart quotes or special dashes

Meals:

""" + '\n\n'.join(meals_text)

    response = claude.messages.create(
        model='claude-sonnet-4-6',
        max_tokens=6000,
        messages=[{'role': 'user', 'content': prompt}]
    )
    data = _clean_json(response.content[0].text)
    return data.get('meals', [])


def generate_recipes_ai(days):
    import concurrent.futures
    claude = get_claude()

    all_meals = []
    for day in days:
        for meal in day['meals']:
            ings = []
            for ing in meal['ingredients']:
                qty_g = ing['qty_g']
                qs = f"{int(qty_g)}g" if qty_g == int(qty_g) else f"{qty_g}g"
                ings.append(f"  - {qs} {ing['food']}")
            all_meals.append((day['day_num'], meal['meal_label'], '\n'.join(ings)))

    BATCH_SIZE = 6
    batches = [all_meals[i:i + BATCH_SIZE] for i in range(0, len(all_meals), BATCH_SIZE)]

    def run_batch(args):
        idx, batch = args
        try:
            results = _generate_batch(claude, batch, idx + 1)
            return results
        except Exception as e:
            print(f"[WARN] Batch {idx + 1} failed: {e}")
            return [{
                'day': day_num,
                'meal_label': meal_label,
                'dish_name': meal_label,
                'steps': ['Prepare ingredients as listed and enjoy.'],
                'unclear': False,
                'unclear_reason': '',
            } for day_num, meal_label, _ in batch]

    with concurrent.futures.ThreadPoolExecutor(max_workers=len(batches)) as executor:
        batch_results = list(executor.map(run_batch, enumerate(batches)))

    ai_results = []
    for results in batch_results:
        ai_results.extend(results)

    ai_map = {(item['day'], item['meal_label']): item for item in ai_results}

    unclear_meals = []
    for day in days:
        for meal in day['meals']:
            key = (day['day_num'], meal['meal_label'])
            if key in ai_map:
                ai = ai_map[key]
                meal['dish_name'] = ai.get('dish_name', meal['meal_label'])
                meal['recipe_steps'] = ai.get('steps', [])
                if ai.get('unclear'):
                    unclear_meals.append({
                        'day': day['day_num'],
                        'meal_label': meal['meal_label'],
                        'unclear_reason': ai.get('unclear_reason', ''),
                        'ingredients': [{'food': i['food'], 'qty_g': i['qty_g']} for i in meal['ingredients']],
                    })

    return days, unclear_meals


# ---------------------------------------------------------------------------
# PDF Generation
# ---------------------------------------------------------------------------
PW, PH = A4
LM, RM = 50.0, 545.28
CW = RM - LM
BLACK    = (0.08, 0.08, 0.08)
WHITE    = (1.0,  1.0,  1.0)
OFFWHITE = (0.96, 0.96, 0.96)
MID_GREY = (0.45, 0.45, 0.45)
LIGHT_GREY = (0.65, 0.65, 0.65)
DIVIDER  = (0.80, 0.80, 0.80)
DARK_CARD= (0.12, 0.12, 0.12)
HB  = 'Inter-Bold'      if INTER_AVAILABLE else 'Helvetica-Bold'
HXB = 'Inter-ExtraBold' if INTER_AVAILABLE else 'Helvetica-Bold'
H   = 'Inter'           if INTER_AVAILABLE else 'Helvetica'

def pdf_y(t): return PH - t
def tw(t, f, s): return stringWidth(t, f, s)

def draw_text(c, x, top_y, text, font, size, rgb):
    c.saveState()
    c.setFillColorRGB(*rgb)
    c.setFont(font, size)
    c.drawString(x, pdf_y(top_y + size), text)
    c.restoreState()

def draw_centred(c, cx, top_y, text, font, size, rgb):
    draw_text(c, cx - tw(text, font, size)/2, top_y, text, font, size, rgb)

def rrect(c, x, bot, w, h, r, fill=None):
    c.saveState()
    c.setLineWidth(0)
    if fill: c.setFillColorRGB(*fill)
    p = c.beginPath()
    p.moveTo(x+r, bot); p.lineTo(x+w-r, bot)
    p.curveTo(x+w, bot, x+w, bot, x+w, bot+r)
    p.lineTo(x+w, bot+h-r)
    p.curveTo(x+w, bot+h, x+w, bot+h, x+w-r, bot+h)
    p.lineTo(x+r, bot+h)
    p.curveTo(x, bot+h, x, bot+h, x, bot+h-r)
    p.lineTo(x, bot+r)
    p.curveTo(x, bot, x, bot, x+r, bot)
    p.close()
    c.drawPath(p, fill=1, stroke=0)
    c.restoreState()

def wrap(text, font, size, max_w):
    words = text.split()
    lines, cur = [], ''
    for word in words:
        test = (cur + ' ' + word).strip()
        if tw(test, font, size) <= max_w: cur = test
        else:
            if cur: lines.append(cur)
            cur = word
    if cur: lines.append(cur)
    return lines

FRUIT_NOTE = ("Note: Any fruit included in today's meals can be eaten as a snack at any point "
              "throughout the day -- it pairs particularly well with Greek yoghurt. "
              "This won't impact your results.")

def draw_cover(c, client_name, logo_b64):
    c.setFillColorRGB(*BLACK)
    c.rect(0, 0, PW, PH, fill=1, stroke=0)

    sz = 100
    logo_top = 280
    logo_bot = logo_top + sz

    if logo_b64:
        try:
            img_data = base64.b64decode(logo_b64)
            with tempfile.NamedTemporaryFile(suffix='.png', delete=False) as tmp:
                tmp.write(img_data); tmp_path = tmp.name
            c.drawImage(tmp_path, PW/2 - sz/2, pdf_y(logo_bot), width=sz, height=sz, mask='auto')
            os.unlink(tmp_path)
        except: pass

    draw_centred(c, PW/2, logo_bot + 14, 'PROJECT GAIN', HB, 9, (0.45, 0.45, 0.45))

    rule_y = logo_bot + 38
    c.saveState()
    c.setStrokeColorRGB(0.25, 0.25, 0.25); c.setLineWidth(0.5)
    c.line(LM + 80, pdf_y(rule_y), RM - 80, pdf_y(rule_y))
    c.restoreState()

    draw_centred(c, PW/2, rule_y + 14, client_name, HXB, 30, WHITE)
    draw_centred(c, PW/2, rule_y + 62, 'Personalised Meal Plan + Recipe Guide', H, 11, (0.45, 0.45, 0.45))
    draw_centred(c, PW/2, PH - 50, 'projectgainofficial.com', H, 8, (0.3, 0.3, 0.3))


def draw_day_header(c, top_y, day_num, kcal, prot, fat, carb, fruit_note):
    box_h = 96.0
    rrect(c, LM, pdf_y(top_y+box_h), CW, box_h, r=6, fill=BLACK)
    draw_centred(c, PW/2, top_y+14, f'Day {day_num}', HXB, 22, WHITE)
    pill_w, pill_h, pill_top = 360.0, 26.0, top_y+44
    rrect(c, PW/2-pill_w/2, pdf_y(pill_top+pill_h), pill_w, pill_h, r=5, fill=(0.2, 0.2, 0.2))
    draw_centred(c, PW/2, pill_top+8, f'{kcal:,} kcal  .  {prot}g Protein  .  {fat}g Fats  .  {carb}g Carbs', HB, 8.5, WHITE)
    y = top_y + box_h + 12
    if fruit_note:
        lines = wrap(FRUIT_NOTE, H, 8.5, CW-20)
        note_h = len(lines)*12+14
        rrect(c, LM, pdf_y(y+note_h), CW, note_h, r=4, fill=OFFWHITE)
        rrect(c, LM, pdf_y(y+note_h), 3, note_h, r=2, fill=DARK_CARD)
        ny = y+7
        for line in lines:
            draw_text(c, LM+12, ny, line, H, 8.5, MID_GREY); ny += 12
        y += note_h+10
    return y


def draw_meal_card(c, top_y, meal_label, dish_name, kcal, prot, fat, carb, ingredients, steps):
    hdr_h = 44.0
    rrect(c, LM, pdf_y(top_y+hdr_h), CW, hdr_h, r=6, fill=DARK_CARD)
    bw, bh, bx, bt = 52.0, 16.0, LM+10, top_y+14
    rrect(c, bx, pdf_y(bt+bh), bw, bh, r=3, fill=(0.28, 0.28, 0.28))
    draw_centred(c, bx+bw/2, bt+4, meal_label.upper(), HB, 7, WHITE)
    if dish_name:
        max_dish_w = RM - (bx + bw + 20)
        dish_display = dish_name
        while dish_display and tw(dish_display, HB, 13) > max_dish_w:
            dish_display = dish_display[:-1]
        if dish_display != dish_name: dish_display += '...'
        draw_text(c, bx+bw+12, top_y+15, dish_display, HXB, 13, WHITE)

    y = top_y + hdr_h
    macro_h = 38.0
    rrect(c, LM, pdf_y(y+macro_h), CW, macro_h, r=0, fill=OFFWHITE)
    divs = [LM+CW*0.25, LM+CW*0.5, LM+CW*0.75]
    centres = [LM+CW*0.125, LM+CW*0.375, LM+CW*0.625, LM+CW*0.875]
    c.saveState()
    c.setStrokeColorRGB(*DIVIDER); c.setLineWidth(0.4)
    for dx in divs: c.line(dx, pdf_y(y+8), dx, pdf_y(y+30))
    c.restoreState()
    for cx, val, lab in zip(centres, [f'{kcal} kcal', f'{prot}g', f'{fat}g', f'{carb}g'], ['Calories','Protein','Fats','Carbs']):
        draw_centred(c, cx, y+7, val, HB, 11, BLACK)
        draw_centred(c, cx, y+20, lab, H, 7, MID_GREY)
    y += macro_h + 18

    draw_text(c, LM, y, 'INGREDIENTS', HB, 8.5, BLACK)
    y += 14

    for ing in ingredients:
        qty_g = ing['qty_g']
        qs = f"{int(qty_g)}g" if qty_g == int(qty_g) else f"{qty_g}g"
        food = ing['food']
        note = get_ingredient_note(food)

        # Always draw exact gram weight in bold
        draw_text(c, LM+8, y, qs, HB, 9.5, BLACK)
        x_after_qty = LM+8+tw(qs, HB, 9.5)+5.5

        # Draw food name in regular weight
        draw_text(c, x_after_qty, y, food, H, 9.5, BLACK)
        x_after_food = x_after_qty + tw(food, H, 9.5)

        # Draw note (raw weight / uncooked weight) in light grey if present
        if note:
            draw_text(c, x_after_food + 5, y, note, H, 8.5, LIGHT_GREY)

        y += 14

    y += 6
    c.saveState(); c.setStrokeColorRGB(*DIVIDER); c.setLineWidth(0.4)
    c.line(LM, pdf_y(y), RM, pdf_y(y)); c.restoreState()
    y += 8

    if steps:
        draw_text(c, LM, y, 'METHOD', HB, 8.5, BLACK)
        y += 14
        max_w = RM - (LM+24)
        for i, step in enumerate(steps, 1):
            draw_text(c, LM+8, y, f'{i}.', HB, 9.5, BLACK)
            lines = wrap(step, H, 9.5, max_w)
            for line in lines:
                draw_text(c, LM+24, y, line, H, 9.5, BLACK); y += 13
            y += 2
    y += 12
    return y


def draw_shopping_list(c, shopping_list):
    ROW_H = 15

    def draw_header(y):
        draw_text(c, LM, y, shopping_list['name'].upper(), HXB, 18, BLACK)
        y += 28
        note_text = f"* {shopping_list['note'].capitalize()}"
        draw_text(c, LM, y, note_text, H, 9, MID_GREY)
        y += 16
        c.saveState(); c.setStrokeColorRGB(*DIVIDER); c.setLineWidth(0.4)
        c.line(LM, pdf_y(y), RM, pdf_y(y)); c.restoreState()
        y += 12
        draw_text(c, LM, y, 'INGREDIENT', HB, 8, MID_GREY)
        draw_text(c, RM - 60, y, 'QUANTITY', HB, 8, MID_GREY)
        y += 12
        return y

    y = 48.0
    y = draw_header(y)

    for i, item in enumerate(shopping_list['items']):
        if y + ROW_H > PH - 40:
            c.showPage()
            y = 48.0
        if i % 2 == 0:
            rrect(c, LM, pdf_y(y + ROW_H), CW, ROW_H, r=0, fill=OFFWHITE)
        draw_text(c, LM + 8, y + 2, item['food'], H, 9.5, BLACK)
        qty_w = tw(item['qty'], HB, 9.5)
        draw_text(c, RM - qty_w - 8, y + 2, item['qty'], HB, 9.5, BLACK)
        y += ROW_H


def generate_pdf_doc(client_name, days, logo_b64, shopping_lists=None):
    buf = io.BytesIO()
    c = canvas.Canvas(buf, pagesize=A4)
    c.setTitle(f"{client_name} -- Meal Plan + Recipe Guide")
    c.setAuthor("Project GAIN")
    draw_cover(c, client_name, logo_b64)
    c.showPage()
    any_fruit = has_fruit(days)
    first_day = True
    y = 48.0

    for day in days:
        hdr_h = 96 + (50 if any_fruit and day_has_fruit(day) else 0) + 12

        if not first_day:
            if y + hdr_h + 150 > PH - 40:
                c.showPage()
                y = 48.0
            else:
                y += 28

        first_day = False

        y = draw_day_header(c, y, day['day_num'],
                            day['total_kcal'], day['total_prot'], day['total_fat'], day['total_carb'],
                            fruit_note=any_fruit and day_has_fruit(day))
        y += 14

        for meal in day['meals']:
            steps = meal.get('recipe_steps', [])
            n_step_lines = sum(len(wrap(s, H, 9.5, RM-(LM+24))) for s in steps)
            card_h = 44+38+18+14+(len(meal['ingredients'])*14)+14+14+(n_step_lines*13)+(len(steps)*2)+12
            if y + card_h > PH - 40:
                c.showPage(); y = 48.0
            y = draw_meal_card(c, y, meal['meal_label'], meal.get('dish_name',''),
                               meal['kcal'], meal['prot'], meal['fat'], meal['carb'],
                               meal['ingredients'], steps)
            y += 14

    if shopping_lists:
        c.showPage()
        for i, sl in enumerate(shopping_lists):
            if i > 0:
                c.showPage()
            draw_shopping_list(c, sl)

    c.save(); buf.seek(0)
    return buf.read()


# ---------------------------------------------------------------------------
# API Routes
# ---------------------------------------------------------------------------
@app.route('/health')
def health():
    return jsonify({'status': 'ok'})


@app.route('/upload', methods=['POST'])
def upload():
    if 'file' not in request.files:
        return jsonify({'error': 'No file uploaded'}), 400
    f = request.files['file']
    if not f.filename.endswith(('.xlsx', '.xls')):
        return jsonify({'error': 'Please upload an Excel file (.xlsx)'}), 400
    file_bytes = f.read()
    try:
        client_name, days, shopping_lists = parse_excel(file_bytes, f.filename)
    except Exception as e:
        return jsonify({'error': f'Could not parse file: {str(e)}'}), 400

    ambiguous = find_ambiguous(days)
    ambiguous = lookup_size_suggestions(ambiguous)

    import uuid
    sid = str(uuid.uuid4())
    _sessions[sid] = {'client_name': client_name, 'days': days, 'shopping_lists': shopping_lists}
    return jsonify({'session_id': sid, 'client_name': client_name, 'day_count': len(days), 'ambiguous': ambiguous})


@app.route('/clarify', methods=['POST'])
def clarify():
    data = request.get_json()
    sid = data.get('session_id')
    if sid not in _sessions:
        return jsonify({'error': 'Session expired. Please re-upload.'}), 400

    sess = _sessions[sid]

    try:
        days, unclear_meals = generate_recipes_ai(sess['days'])
    except Exception as e:
        import traceback
        return jsonify({'error': f'Recipe generation failed: {str(e)}', 'detail': traceback.format_exc()}), 500

    sess['days'] = days

    review_days = []
    for day in days:
        review_days.append({
            'day_num': day['day_num'],
            'total_kcal': day['total_kcal'], 'total_prot': day['total_prot'],
            'total_fat': day['total_fat'],   'total_carb': day['total_carb'],
            'meals': [{
                'meal_label': m['meal_label'],
                'dish_name':  m.get('dish_name', ''),
                'kcal': m['kcal'], 'prot': m['prot'], 'fat': m['fat'], 'carb': m['carb'],
                'ingredients': [{'food': i['food'], 'qty_g': i['qty_g'],
                                 'excel_qty_g': i.get('excel_qty_g', i['qty_g'])} for i in m['ingredients']],
                'recipe_steps': m.get('recipe_steps', []),
                'needs_clarification': m.get('needs_clarification', False),
            } for m in day['meals']],
        })

    return jsonify({'session_id': sid, 'review_days': review_days, 'unclear_meals': unclear_meals})


@app.route('/generate-pdf', methods=['POST'])
def generate_pdf_route():
    data = request.get_json()
    sid = data.get('session_id')
    approved_days = data.get('approved_days', [])
    logo_b64 = data.get('logo_b64', '')

    if sid not in _sessions:
        return jsonify({'error': 'Session expired. Please re-upload.'}), 400

    sess = _sessions[sid]
    days = sess['days']
    client_name = sess['client_name']

    qty_issues = []
    for i, day in enumerate(days):
        if i >= len(approved_days): continue
        adv = approved_days[i]
        for j, meal in enumerate(day['meals']):
            if j >= len(adv.get('meals', [])): continue
            am = adv['meals'][j]
            meal['dish_name']    = am.get('dish_name', meal.get('dish_name', ''))
            meal['recipe_steps'] = am.get('recipe_steps', meal.get('recipe_steps', []))
            for k, ing in enumerate(meal['ingredients']):
                excel_qty = ing.get('excel_qty_g', ing['qty_g'])
                if k < len(am.get('ingredients', [])):
                    sub_qty = float(am['ingredients'][k].get('qty_g', ing['qty_g']))
                    if abs(excel_qty - sub_qty) > 0.01:
                        qty_issues.append({'day': day['day_num'], 'meal': meal['meal_label'],
                                           'food': ing['food'], 'excel_qty': excel_qty, 'submitted_qty': sub_qty})

    if qty_issues:
        return jsonify({'qty_issues': qty_issues, 'error': 'Quantity mismatch detected -- please review.'}), 422

    shopping_lists = sess.get('shopping_lists', [])
    try:
        pdf_bytes = generate_pdf_doc(client_name, days, logo_b64, shopping_lists)
    except Exception as e:
        import traceback
        return jsonify({'error': f'PDF failed: {str(e)}', 'detail': traceback.format_exc()}), 500

    del _sessions[sid]
    return jsonify({'pdf_b64': base64.b64encode(pdf_bytes).decode(),
                    'filename': f"{client_name.replace(' ', '_')}_Meal_Plan.pdf"})


if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    app.run(host='0.0.0.0', port=port, debug=False)
